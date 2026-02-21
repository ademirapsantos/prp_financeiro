from flask import Blueprint, render_template, request, redirect, url_for, send_file, current_app, jsonify
import requests
from flask_login import login_required, current_user
import json
import io
import os
import math
import markdown
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import Notificacao, UpdateLog, Configuracao, db, ContaContabil, Titulo, LivroDiario, PartidaDiario, TipoConta, StatusTitulo, TipoTitulo, CartaoCredito, FaturaCartao
from .version import __version__, __build__
from datetime import datetime, timedelta
from sqlalchemy import func

main_bp = Blueprint('main', __name__)

@main_bp.before_app_request
def check_maintenance():
    if Configuracao.is_maintenance():
        # Permitir apenas rotas críticas, estáticos e endpoints de controle de update
        allowed_paths = [
            '/api/version', 
            '/api/system/latest', 
            '/api/system/update/status',
            '/api/system/update/finalize-token',
            '/api/system/maintenance/off',
            '/api/system/maintenance/off-token', # Emergência
            '/health', 
            '/static/', 
            '/login', 
            '/logout'
        ]
        if not any(request.path.startswith(p) for p in allowed_paths):
            if request.path.startswith('/api/'):
                return jsonify({"status": "maintenance", "message": "Sistema em manutenção"}), 503
            return render_template('manutencao.html'), 503

@main_bp.route('/health')
def health():
    if Configuracao.is_maintenance():
        return jsonify({
            "status": "maintenance", 
            "version": __version__,
            "update_in_progress": Configuracao.get_valor('UPDATE_IN_PROGRESS') == 'true'
        }), 503
    return jsonify({
        "status": "healthy", 
        "version": __version__,
        "build": __build__
    })

@main_bp.route('/api/version')
def api_version():
    return {
        "version": __version__,
        "build": __build__
    }

@main_bp.route('/')
def dashboard():
    from .models import TransacaoFinanceira, TipoTitulo, StatusTitulo, TipoTransacao, CartaoCredito, FaturaCartao, LivroDiario, PartidaDiario, ContaContabil
    from sqlalchemy import extract, or_
    import calendar

    # 1. Filtros e Datas
    ano_atual = datetime.utcnow().year
    ano = request.args.get('ano', ano_atual, type=int)
    mes_filtro = request.args.get('mes', type=int)

    # Determinar data limite para cálculos de saldo (Disponível/Ativos)
    if mes_filtro:
        ultimo_dia = calendar.monthrange(ano, mes_filtro)[1]
        data_limite = datetime(ano, mes_filtro, ultimo_dia, 23, 59, 59)
    else:
        data_limite = datetime(ano, 12, 31, 23, 59, 59)

    # 2. Calcular Disponível (Até a Data Limite)
    contas_disponivel_ids = db.session.query(ContaContabil.id).filter(
        or_(ContaContabil.codigo.like('1.1%'), ContaContabil.codigo.like('1.2%')),
        ~ContaContabil.codigo.like('1.1.05%'),
        ~ContaContabil.codigo.like('1.5%')
    ).subquery()
    
    res_disp = db.session.query(
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
    ).join(LivroDiario, PartidaDiario.diario_id == LivroDiario.id)\
     .filter(PartidaDiario.conta_id.in_(contas_disponivel_ids))\
     .filter(LivroDiario.data <= data_limite).first()
    disponivel = (res_disp.debitos or 0) - (res_disp.creditos or 0) if res_disp else 0

    # 3. Métricas de Cartão (Sincronizadas com o Filtro)
    if mes_filtro:
        # Se houver mês, mostramos o comportamento DO MÊS
        # 3.1 Total Gasto no Mês (Fatura)
        cartao_limite_disponivel = db.session.query(func.sum(FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano, 
                    extract('month', FaturaCartao.data_vencimento) == mes_filtro).scalar() or 0
        cartao_label_limite = "Gasto no Mês"
        
        # 3.2 Saldo da Fatura Selecionada (Pendente de Pagamento)
        total_ciclo_aberto = db.session.query(func.sum(FaturaCartao.total - func.coalesce(FaturaCartao.total_pago, 0)))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano, 
                    extract('month', FaturaCartao.data_vencimento) == mes_filtro).scalar() or 0
        cartao_label_ciclo = "Fatura Pendente"
    else:
        # Sem mês: mostramos o LIMITE ATUAL e as FATURAS EM ABERTO ATUAIS
        cartao_limite_disponivel = db.session.query(func.sum(CartaoCredito.limite_disponivel)).scalar() or 0
        cartao_label_limite = "Limite Disponível"
        
        total_ciclo_aberto = db.session.query(func.sum(FaturaCartao.total - func.coalesce(FaturaCartao.total_pago, 0)))\
            .filter(FaturaCartao.status == 'aberta').scalar() or 0
        cartao_label_ciclo = "Fatura Atual"

    # 4. Totais (A Receber / A Pagar) - Filtrados por Ano e Mês
    def get_total_titulos(tipo, mes=None):
        query = db.session.query(func.sum(Titulo.valor)).filter(
            Titulo.tipo == tipo,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano
        )
        if mes:
            query = query.filter(extract('month', Titulo.data_vencimento) == mes)
        return query.scalar() or 0

    total_a_receber = get_total_titulos(TipoTitulo.RECEBER.value, mes_filtro)
    
    # A Pagar Original (Títulos)
    total_a_pagar_titulos = get_total_titulos(TipoTitulo.PAGAR.value, mes_filtro)
    
    # A Pagar Faturas Fechadas (Mês Selecionado ou Ano Inteiro)
    query_faturas_fechadas = db.session.query(func.sum(FaturaCartao.total - func.coalesce(FaturaCartao.total_pago, 0)))\
        .filter(FaturaCartao.status == 'fechada')\
        .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
        .filter(extract('year', FaturaCartao.data_vencimento) == ano)
    
    if mes_filtro:
        query_faturas_fechadas = query_faturas_fechadas.filter(extract('month', FaturaCartao.data_vencimento) == mes_filtro)
    
    total_faturas_fechadas = query_faturas_fechadas.scalar() or 0
    total_a_pagar = total_a_pagar_titulos + total_faturas_fechadas

    # 5. Patrimônio Líquido (Ativos na data - Dívidas na data)
    res_ativos_total = db.session.query(
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
    ).join(ContaContabil).join(LivroDiario, PartidaDiario.diario_id == LivroDiario.id)\
     .filter(ContaContabil.codigo.like('1%'), ~ContaContabil.codigo.like('1.5%'))\
     .filter(LivroDiario.data <= data_limite).first()
    
    total_ativos_regime_caixa = (res_ativos_total.debitos or 0) - (res_ativos_total.creditos or 0)
    patrimonio_liquido = total_ativos_regime_caixa - total_a_pagar

    # 6. Dados para o Gráfico (Dinâmico: 12 meses ou Mês Selecionado)
    chart_data = {'recebido': [], 'pago': [], 'a_receber': [], 'a_pagar': []}
    chart_labels = []
    
    meses_nomes = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    iteracao_meses = [mes_filtro] if mes_filtro else range(1, 13)

    for m in iteracao_meses:
        chart_labels.append(meses_nomes[m-1])
        
        # Recebido / Pago
        res_trans = db.session.query(
            func.sum(TransacaoFinanceira.valor).filter(TransacaoFinanceira.tipo == TipoTransacao.RECEBIMENTO.value).label('recebido'),
            func.sum(TransacaoFinanceira.valor).filter(TransacaoFinanceira.tipo == TipoTransacao.PAGAMENTO.value).label('pago')
        ).outerjoin(Titulo).filter(
            extract('year', TransacaoFinanceira.data) == ano,
            extract('month', TransacaoFinanceira.data) == m,
            or_(Titulo.status != StatusTitulo.CANCELADO.value, Titulo.id == None)
        ).first()
        
        # A Receber
        a_receber_m = db.session.query(func.sum(Titulo.valor)).filter(
            Titulo.tipo == TipoTitulo.RECEBER.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == m
        ).scalar() or 0
        
        # A Pagar (Títulos + Faturas Fechadas)
        a_pagar_titulos_m = db.session.query(func.sum(Titulo.valor)).filter(
            Titulo.tipo == TipoTitulo.PAGAR.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == m
        ).scalar() or 0
        
        a_pagar_faturas_m = db.session.query(func.sum(FaturaCartao.total - func.coalesce(FaturaCartao.total_pago, 0)))\
            .filter(FaturaCartao.status == 'fechada')\
            .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano, extract('month', FaturaCartao.data_vencimento) == m)\
            .scalar() or 0
            
        a_pagar_total_m = a_pagar_titulos_m + a_pagar_faturas_m
        
        chart_data['recebido'].append(float(res_trans.recebido or 0))
        chart_data['pago'].append(float(res_trans.pago or 0))
        chart_data['a_receber'].append(float(a_receber_m))
        chart_data['a_pagar'].append(float(a_pagar_total_m))

    anos_disponiveis = list(range(2025, 2151))

    return render_template('dashboard.html', 
                         patrimonio_liquido=patrimonio_liquido,
                         disponivel=disponivel,
                         total_a_receber=total_a_receber,
                         total_a_pagar=total_a_pagar,
                         cartao_limite_disponivel=cartao_limite_disponivel,
                         cartao_label_limite=cartao_label_limite,
                         total_ciclo_aberto=total_ciclo_aberto,
                         cartao_label_ciclo=cartao_label_ciclo,
                         chart_data=chart_data,
                         chart_labels=chart_labels,
                         ano_selecionado=ano,
                         mes_selecionado=mes_filtro,
                         anos_disponiveis=anos_disponiveis)

@main_bp.route('/contabilidade/diario')
def diario():
    from sqlalchemy.orm import joinedload
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    
    query = LivroDiario.query.options(
        joinedload(LivroDiario.partidas).joinedload(PartidaDiario.conta)
    )
    
    if data_inicio:
        try:
            di = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(LivroDiario.data >= di)
        except ValueError:
            pass
            
    if data_fim:
        try:
            df = datetime.strptime(data_fim, '%Y-%m-%d')
            # Ajustar para o fim do dia
            df = df.replace(hour=23, minute=59, second=59)
            query = query.filter(LivroDiario.data <= df)
        except ValueError:
            pass
            
    pagination = query.order_by(LivroDiario.data.desc(), LivroDiario.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    diarios = pagination.items
    
    return render_template('contabilidade/diario.html', 
                           diarios=diarios, 
                           pagination=pagination,
                           data_inicio=data_inicio,
                           data_fim=data_fim)

@main_bp.route('/contabilidade/diario/exportar')
def exportar_diario():
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    query = LivroDiario.query
    
    if data_inicio_str:
        try:
            di = datetime.strptime(data_inicio_str, '%Y-%m-%d')
            query = query.filter(LivroDiario.data >= di)
        except ValueError:
            di = None
    else:
        di = None
            
    if data_fim_str:
        try:
            df = datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(LivroDiario.data <= df)
        except ValueError:
            df = None
    else:
        df = None
            
    diarios = query.order_by(LivroDiario.data.asc(), LivroDiario.id.asc()).all()

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Livro Diario"

    # Cabeçalhos
    headers = ["Data", "ID", "Histórico", "Conta", "Débito", "Crédito"]
    ws.append(headers)
    
    # Estilo Cabeçalho
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for d in diarios:
        # Linha do Lançamento (Cabeçalho do lançamento no Excel)
        ws.append([d.data.strftime('%d/%m/%Y'), d.id, d.historico, "", "", ""])
        current_row = ws.max_row
        for cell in ws[current_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        
        # Partidas
        for p in d.partidas:
            debito = p.valor if p.tipo == 'D' else 0
            credito = p.valor if p.tipo == 'C' else 0
            
            row = [
                "", # Data vazia para os itens
                "", # ID vazio
                "", # Historico vazio
                f"{p.conta.codigo} - {p.conta.nome}",
                debito if debito > 0 else "",
                credito if credito > 0 else ""
            ]
            ws.append(row)
            
            # Formatação de valores
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right")

    # Ajustar largura das colunas
    column_widths = [12, 8, 40, 40, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"diario_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def get_balancete_results(data_inicio, data_fim):
    from .models import PartidaDiario, LivroDiario, ContaContabil
    from sqlalchemy import func
    
    # 1. Buscar todas as contas
    contas = ContaContabil.query.order_by(ContaContabil.codigo).all()
    
    # 2. Buscar somatórios de todas as partidas para evitar N+1
    res_ant_all = db.session.query(
        PartidaDiario.conta_id,
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
    ).join(LivroDiario).filter(
        LivroDiario.data < data_inicio
    ).group_by(PartidaDiario.conta_id).all()
    
    map_ant = {r.conta_id: (r.debitos or 0, r.creditos or 0) for r in res_ant_all}
    
    res_per_all = db.session.query(
        PartidaDiario.conta_id,
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
        func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
    ).join(LivroDiario).filter(
        LivroDiario.data >= data_inicio,
        LivroDiario.data <= data_fim
    ).group_by(PartidaDiario.conta_id).all()
    
    map_per = {r.conta_id: (r.debitos or 0, r.creditos or 0) for r in res_per_all}
    
    balancete_data = []
    
    for conta in contas:
        prefixo = conta.codigo
        descendentes = [c for c in contas if c.codigo == prefixo or c.codigo.startswith(prefixo + '.')]
        
        deb_ant_total = 0
        cre_ant_total = 0
        deb_per_total = 0
        cre_per_total = 0
        
        for d in descendentes:
            ant = map_ant.get(d.id, (0, 0))
            per = map_per.get(d.id, (0, 0))
            deb_ant_total += ant[0]
            cre_ant_total += ant[1]
            deb_per_total += per[0]
            cre_per_total += per[1]
            
        if conta.natureza == 'Devedora':
            saldo_anterior_net = deb_ant_total - cre_ant_total
            saldo_atual_net = saldo_anterior_net + deb_per_total - cre_per_total
            
            saldo_ant_deb = saldo_anterior_net if saldo_anterior_net > 0 else 0
            saldo_ant_cre = abs(saldo_anterior_net) if saldo_anterior_net < 0 else 0
            saldo_atu_deb = saldo_atual_net if saldo_atual_net > 0 else 0
            saldo_atu_cre = abs(saldo_atual_net) if saldo_atual_net < 0 else 0
        else:
            # Natureza Credora: Saldo positivo é crédito
            saldo_anterior_net = cre_ant_total - deb_ant_total
            saldo_atual_net = saldo_anterior_net + cre_per_total - deb_per_total
            
            saldo_ant_cre = saldo_anterior_net if saldo_anterior_net > 0 else 0
            saldo_ant_deb = abs(saldo_anterior_net) if saldo_anterior_net < 0 else 0
            saldo_atu_cre = saldo_atual_net if saldo_atual_net > 0 else 0
            saldo_atu_deb = abs(saldo_atual_net) if saldo_atual_net < 0 else 0

        nivel = len(conta.codigo.split('.'))
        tem_valor = (abs(deb_ant_total) + abs(cre_ant_total) + abs(deb_per_total) + abs(cre_per_total)) > 0.001
        
        if tem_valor or nivel == 1:
            balancete_data.append({
                'codigo': conta.codigo,
                'nome': conta.nome,
                'natureza': conta.natureza,
                'tipo': conta.tipo,
                'saldo_ant_deb': saldo_ant_deb,
                'saldo_ant_cre': saldo_ant_cre,
                'debitos': deb_per_total,
                'creditos': cre_per_total,
                'saldo_atu_deb': saldo_atu_deb,
                'saldo_atu_cre': saldo_atu_cre,
                'nivel': nivel,
                'analitica': conta.is_analitica
            })
    return balancete_data

@main_bp.route('/contabilidade/balancete')
def balancete():
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    hoje = datetime.utcnow()
    if not data_inicio_str:
        data_inicio = hoje.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
        
    if not data_fim_str:
        data_fim = hoje.replace(hour=23, minute=59, second=59, microsecond=999999)
    else:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)

    balancete_data = get_balancete_results(data_inicio, data_fim)

    # Calcular totais das colunas (Soma de Nível 1 para evitar duplicidade)
    totais = {
        'saldo_ant_deb': 0,
        'saldo_ant_cre': 0,
        'debitos': 0,
        'creditos': 0,
        'saldo_atu_deb': 0,
        'saldo_atu_cre': 0
    }
    
    for item in balancete_data:
        if item['nivel'] == 1:
            totais['saldo_ant_deb'] += item['saldo_ant_deb']
            totais['saldo_ant_cre'] += item['saldo_ant_cre']
            totais['debitos'] += item['debitos']
            totais['creditos'] += item['creditos']
            totais['saldo_atu_deb'] += item['saldo_atu_deb']
            totais['saldo_atu_cre'] += item['saldo_atu_cre']

    # Para compatibilidade com alertas e indicadores, saldo_anterior e saldo_atual
    # agora representam o volume total (D ou C)
    totais['saldo_anterior'] = max(totais['saldo_ant_deb'], totais['saldo_ant_cre'])
    totais['saldo_atual'] = max(totais['saldo_atu_deb'], totais['saldo_atu_cre'])

    return render_template('contabilidade/balancete.html',
                           balancete=balancete_data,
                           totais=totais,
                           data_inicio=data_inicio.strftime('%Y-%m-%d'),
                           data_fim=data_fim.strftime('%Y-%m-%d'))


@main_bp.route('/contabilidade/balancete/exportar')
def exportar_balancete():
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    
    if data_inicio_str:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
    else:
        data_inicio = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0)
        
    if data_fim_str:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    else:
        data_fim = datetime.utcnow().replace(hour=23, minute=59, second=59)

    balancete_data = get_balancete_results(data_inicio, data_fim)

    # Criar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Balancete"

    # Cabeçalhos
    headers = ["Código", "Conta", "S. Anterior Devedor", "S. Anterior Credor", "Débitos (Período)", "Créditos (Período)", "S. Atual Devedor", "S. Atual Credor"]
    ws.append(headers)
    
    # Estilo Cabeçalho
    header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Dados
    for item in balancete_data:
        row = [
            item['codigo'],
            item['nome'],
            item['saldo_ant_deb'],
            item['saldo_ant_cre'],
            item['debitos'],
            item['creditos'],
            item['saldo_atu_deb'],
            item['saldo_atu_cre']
        ]
        ws.append(row)
        
        # Formatação por nível
        current_row = ws.max_row
        if item['nivel'] == 1:
            for cell in ws[current_row]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        elif item['nivel'] == 2:
            for cell in ws[current_row]:
                cell.font = Font(bold=True)

        # Alinhamento das colunas de valor (3 a 8)
        for col in range(3, 9):
            ws.cell(row=current_row, column=col).number_format = '#,##0.00'
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="right")

    # Ajustar largura das colunas
    column_widths = [15, 45, 18, 18, 18, 18, 18, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"balancete_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@main_bp.route('/detalhamento/<int:ano>/<int:mes>/<tipo>')
def detalhamento(ano, mes, tipo):
    from .models import LivroDiario, PartidaDiario, ContaContabil, Titulo, TransacaoFinanceira, TipoTransacao, StatusTitulo, FaturaCartao, Configuracao
    from sqlalchemy import extract, func, or_
    from sqlalchemy.orm import joinedload
    
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    meses_nomes = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", 
                   "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    mes_nome = meses_nomes[mes - 1]
    
    query = Titulo.query.options(joinedload(Titulo.entidade))
    total_query = db.session.query(func.sum(Titulo.valor))
    tipo_label = ""
    faturas_detalhe = []
    
    if tipo == 'recebido':
        tipo_label = "Recebido"
        titulos_ids_query = db.session.query(TransacaoFinanceira.titulo_id).filter(
            TransacaoFinanceira.tipo == TipoTransacao.RECEBIMENTO.value,
            extract('year', TransacaoFinanceira.data) == ano,
            extract('month', TransacaoFinanceira.data) == mes,
            TransacaoFinanceira.titulo_id.isnot(None)
        ).distinct()
        
        query = query.options(
            joinedload(Titulo.transacoes).joinedload(TransacaoFinanceira.ativo)
        ).filter(Titulo.id.in_(titulos_ids_query.subquery()))
        
        total = total_query.filter(Titulo.id.in_(titulos_ids_query.subquery())).scalar() or 0
            
    elif tipo == 'pago':
        tipo_label = "Pago"
        titulos_ids_query = db.session.query(TransacaoFinanceira.titulo_id).filter(
            TransacaoFinanceira.tipo == TipoTransacao.PAGAMENTO.value,
            extract('year', TransacaoFinanceira.data) == ano,
            extract('month', TransacaoFinanceira.data) == mes,
            TransacaoFinanceira.titulo_id.isnot(None)
        ).distinct()
        
        query = query.options(
            joinedload(Titulo.transacoes).joinedload(TransacaoFinanceira.ativo)
        ).filter(Titulo.id.in_(titulos_ids_query.subquery()))
        
        total = total_query.filter(Titulo.id.in_(titulos_ids_query.subquery())).scalar() or 0
            
    elif tipo == 'a_receber':
        tipo_label = "A Receber"
        query = query.filter(
            Titulo.tipo == TipoTitulo.RECEBER.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == mes
        )
        total = total_query.filter(
            Titulo.tipo == TipoTitulo.RECEBER.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == mes
        ).scalar() or 0
        
    elif tipo == 'a_pagar':
        tipo_label = "A Pagar"
        query = query.filter(
            Titulo.tipo == TipoTitulo.PAGAR.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == mes
        )
        total_titulos = total_query.filter(
            Titulo.tipo == TipoTitulo.PAGAR.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano,
            extract('month', Titulo.data_vencimento) == mes
        ).scalar() or 0
        
        # Somar faturas fechadas no mês
        total_faturas = db.session.query(func.sum(FaturaCartao.total - func.coalesce(FaturaCartao.total_pago, 0)))\
            .filter(FaturaCartao.status == 'fechada')\
            .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano, extract('month', FaturaCartao.data_vencimento) == mes)\
            .scalar() or 0
            
        total = total_titulos + total_faturas
        
        faturas_detalhe = FaturaCartao.query.filter(FaturaCartao.status == 'fechada')\
            .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano, extract('month', FaturaCartao.data_vencimento) == mes)\
            .all()
    
    pagination = query.order_by(Titulo.data_vencimento.asc()).paginate(page=page, per_page=per_page, error_out=False)
    titulos = pagination.items
    
    return render_template('detalhamento.html',
                         titulos=titulos,
                         pagination=pagination,
                         total=total,
                         mes=mes, # Adicionado para facilitar links
                         mes_nome=mes_nome,
                         ano=ano,
                         tipo=tipo, # Necessário para os links de paginação
                         tipo_label=tipo_label)

@main_bp.route('/ajuda')
def ajuda():
    import os
    import markdown
    from flask import current_app
    manual_path = os.path.join(current_app.root_path, '..', 'MANUAL_USUARIO.md')
    try:
        with open(manual_path, 'r', encoding='utf-8') as f:
            text = f.read()
            # Converte Markdown para HTML com suporte a tabelas e outras extensões comuns
            html = markdown.markdown(text, extensions=['tables', 'fenced_code', 'nl2br'])
            return render_template('ajuda.html', manual_html=html)
    except FileNotFoundError:
        return "Manual não encontrado.", 404

@main_bp.route('/api/dashboard/drilldown')
def api_drilldown():
    from .models import TipoConta, StatusTitulo, TipoTitulo, CartaoCredito, FaturaCartao, LivroDiario, PartidaDiario, ContaContabil, Titulo
    from sqlalchemy import extract, or_, func
    import math
    import calendar
    
    tipo = request.args.get('tipo')
    ano = request.args.get('ano', datetime.utcnow().year, type=int)
    mes = request.args.get('mes', type=int)
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    # Determinar data limite para cálculos de saldo (Disponível/Ativos)
    if mes:
        ultimo_dia = calendar.monthrange(ano, mes)[1]
        data_limite = datetime(ano, mes, ultimo_dia, 23, 59, 59)
    else:
        # Fim do ano filtrado
        data_limite = datetime(ano, 12, 31, 23, 59, 59)

    data = {
        'title': '',
        'items': [],
        'total': 0
    }

    all_items = []

    if tipo == 'patrimonio':
        # Detalhamento: Ativos (Exceto 1.5) - A Pagar (Regra Refinada)
        data['title'] = 'Composição do Patrimônio Líquido (Ativos - Dívidas)'
        
        # 1. Ativos (Cronológico até data_limite)
        ativos = db.session.query(
            ContaContabil.codigo, 
            ContaContabil.nome,
            func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
            func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
        ).join(PartidaDiario).join(LivroDiario, PartidaDiario.diario_id == LivroDiario.id)\
        .filter(ContaContabil.codigo.like('1%'), ~ContaContabil.codigo.like('1.5%'))\
        .filter(LivroDiario.data <= data_limite)\
        .group_by(ContaContabil.id).all()

        for a in ativos:
            saldo = (a.debitos or 0) - (a.creditos or 0)
            if abs(saldo) > 0.01:
                all_items.append({
                    'label': f"{a.codigo} - {a.nome}",
                    'valor': float(saldo),
                    'tipo': 'Ativo'
                })

        # 2. Títulos a Pagar (No Ano, opcional Mês)
        query_pagar = Titulo.query.filter(
            Titulo.tipo == TipoTitulo.PAGAR.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano
        )
        if mes:
            query_pagar = query_pagar.filter(extract('month', Titulo.data_vencimento) == mes)
        
        titulos = query_pagar.order_by(Titulo.data_vencimento.asc()).all()
        for t in titulos:
            all_items.append({
                'label': f"{t.data_vencimento.strftime('%d/%m/%Y')} - {t.descricao} ({t.entidade.nome})",
                'valor': float(-t.valor),
                'tipo': 'A Pagar (Título)'
            })
            
        # 3. Faturas Fechadas Não Pagas (No Ano, opcional Mês)
        query_faturas = FaturaCartao.query.filter(FaturaCartao.status == 'fechada')\
            .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano)
            
        if mes:
            query_faturas = query_faturas.filter(extract('month', FaturaCartao.data_vencimento) == mes)
            
        faturas = query_faturas.all()
        for f in faturas:
            saldo_f = f.total - (f.total_pago or 0)
            all_items.append({
                'label': f"Fatura {f.cartao.nome} - Venc: {f.data_vencimento.strftime('%d/%m/%Y')}",
                'valor': float(-saldo_f),
                'tipo': 'A Pagar (Cartão Fechado)'
            })

    elif tipo == 'disponivel':
        data['title'] = 'Saldos em Contas Disponíveis'
        
        contas_ids = db.session.query(ContaContabil.id).filter(
            or_(ContaContabil.codigo.like('1.1%'), ContaContabil.codigo.like('1.2%')),
            ~ContaContabil.codigo.like('1.1.05%'),
            ~ContaContabil.codigo.like('1.5%')
        ).subquery()

        saldos = db.session.query(
            ContaContabil.codigo, 
            ContaContabil.nome,
            func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'D').label('debitos'),
            func.sum(PartidaDiario.valor).filter(PartidaDiario.tipo == 'C').label('creditos')
        ).join(PartidaDiario).join(LivroDiario, PartidaDiario.diario_id == LivroDiario.id)\
        .filter(PartidaDiario.conta_id.in_(contas_ids))\
        .filter(LivroDiario.data <= data_limite)\
        .group_by(ContaContabil.id).all()

        for s in saldos:
            saldo = (s.debitos or 0) - (s.creditos or 0)
            if abs(saldo) > 0.01:
                all_items.append({
                    'label': f"{s.codigo} - {s.nome}",
                    'valor': float(saldo)
                })

    elif tipo == 'a_receber':
        data['title'] = f'Títulos a Receber'
        query = Titulo.query.filter(
            Titulo.tipo == TipoTitulo.RECEBER.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano
        )
        if mes:
            query = query.filter(extract('month', Titulo.data_vencimento) == mes)
        
        titulos = query.order_by(Titulo.data_vencimento.asc()).all()
        for t in titulos:
            all_items.append({
                'label': f"{t.data_vencimento.strftime('%d/%m/%Y')} - {t.descricao} ({t.entidade.nome})",
                'valor': float(t.valor)
            })

    elif tipo == 'a_pagar':
        data['title'] = f'Dívidas Totais'
        
        # 1. Títulos a Pagar
        query_titulos = Titulo.query.filter(
            Titulo.tipo == TipoTitulo.PAGAR.value,
            Titulo.status == StatusTitulo.ABERTO.value,
            extract('year', Titulo.data_vencimento) == ano
        )
        if mes:
            query_titulos = query_titulos.filter(extract('month', Titulo.data_vencimento) == mes)
        
        titulos = query_titulos.order_by(Titulo.data_vencimento.asc()).all()
        for t in titulos:
            all_items.append({
                'label': f"{t.data_vencimento.strftime('%d/%m/%Y')} - {t.descricao} ({t.entidade.nome})",
                'valor': float(t.valor),
                'tipo': 'Título'
            })
            
        # 2. Faturas Fechadas Não Pagas
        query_faturas = FaturaCartao.query.filter(FaturaCartao.status == 'fechada')\
            .filter(or_(FaturaCartao.situacao_pagamento != 'paga', FaturaCartao.total_pago < FaturaCartao.total))\
            .filter(extract('year', FaturaCartao.data_vencimento) == ano)
        
        if mes:
            query_faturas = query_faturas.filter(extract('month', FaturaCartao.data_vencimento) == mes)
            
        faturas = query_faturas.all()
        for f in faturas:
            saldo = f.total - (f.total_pago or 0)
            all_items.append({
                'label': f"Fatura {f.cartao.nome} - Venc: {f.data_vencimento.strftime('%d/%m/%Y')}",
                'valor': float(saldo),
                'tipo': 'Cartão (Fatura Fechada)'
            })

    elif tipo == 'cartao_limite_disponivel':
        if mes:
            data['title'] = f'Gasto no Cartão de Crédito ({mes}/{ano})'
            faturas = FaturaCartao.query.filter(
                extract('year', FaturaCartao.data_vencimento) == ano,
                extract('month', FaturaCartao.data_vencimento) == mes
            ).all()
            for f in faturas:
                all_items.append({
                    'label': f"Cartão {f.cartao.nome} - Fatura Venc: {f.data_vencimento.strftime('%d/%m/%Y')}",
                    'valor': float(f.total),
                    'tipo': f'Total Gasto: R$ {f.total:,.2f}'
                })
        else:
            data['title'] = 'Limite Disponível por Cartão (Total Atual)'
            cartoes = CartaoCredito.query.filter_by(ativo=True).all()
            for c in cartoes:
                all_items.append({
                    'label': f"{c.nome} (Limite Total: R$ {c.limite_total:,.2f})",
                    'valor': float(c.limite_disponivel or 0),
                    'tipo': 'Limite Disponível'
                })

    elif tipo == 'cartao_ciclo_aberto':
        if mes:
            data['title'] = f'Fatura Pendente ({mes}/{ano})'
            faturas = FaturaCartao.query.filter(
                extract('year', FaturaCartao.data_vencimento) == ano,
                extract('month', FaturaCartao.data_vencimento) == mes
            ).all()
        else:
            data['title'] = 'Fatura Atual (Ciclo Aberto)'
            faturas = FaturaCartao.query.filter_by(status='aberta').all()
            
        for f in faturas:
            saldo_f = f.total - (f.total_pago or 0)
            if abs(saldo_f) > 0.01:
                all_items.append({
                    'label': f"Fatura {f.cartao.nome} ({f.competencia})",
                    'valor': float(saldo_f),
                    'tipo': f'A Pagar: R$ {saldo_f:,.2f}'
                })

    # Paginação manual da lista consolidada
    total_items = len(all_items)
    total_valor = sum(i['valor'] for i in all_items)
    total_pages = math.ceil(total_items / per_page) if total_items > 0 else 1
    
    start = (page - 1) * per_page
    end = start + per_page
    paginated_items = all_items[start:end]

    return {
        'title': data['title'],
        'items': paginated_items,
        'total': float(total_valor),
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total_items': total_items,
            'total_pages': total_pages,
            'has_prev': page > 1,
            'has_next': page < total_pages
        }
    }

@main_bp.route('/api/contabilidade/parametros', methods=['GET', 'POST'])
def api_contabilidade_parametros():
    from .models import Configuracao, ContaContabil
    
    if request.method == 'POST':
        dados = request.get_json()
        for chave, valor in dados.items():
            Configuracao.set_valor(chave, valor)
        return {"status": "success"}

    # GET: Retornar contas analíticas e valores atuais
    contas = ContaContabil.query.all()
    # Somente contas analíticas (que não possuem subcontas)
    analiticas = [{"id": c.id, "codigo": c.codigo, "nome": c.nome, "tipo": c.tipo} for c in contas if c.is_analitica]
    analiticas.sort(key=lambda x: x['codigo'])

    chaves = [
        'conta_lucro_venda', 'conta_prejuizo_venda',
        'conta_ativo_banco', 'conta_ativo_veiculo', 
        'conta_ativo_imovel', 'conta_ativo_investimento', 'conta_ativo_outros',
        'CONTA_DESCONTO_OBTIDO_ID', 'CONTA_DESCONTO_CONCEDIDO_ID'
    ]
    
    valores = {chave: Configuracao.get_valor(chave) for chave in chaves}
    
    return {
        "contas": analiticas,
        "valores": valores
    }

# --- Módulo de Atualização e Notificações ---

def parse_version(v_str):
    """Converte string de versão (ex: '1.4.1') em tupla de inteiros para comparação."""
    try:
        return tuple(map(int, (v_str.strip().replace('v', '').split('.'))))
    except:
        return (0, 0, 0)

@main_bp.route('/api/system/latest')
def system_latest():
    """
    Retorna a versão mais recente do sistema disponível para o ambiente atual.
    Suporta manifestos com chaves 'version' ou 'latest_version'.
    """
    repo = os.getenv('GITHUB_REPO', 'ademirapsantos/prp_financeiro')
    env = os.getenv('ENVIRONMENT', 'dev')
    flask_env = os.getenv('FLASK_ENV', 'production')
    
    manifest_file = os.getenv('MANIFEST_FILE')
    manifest_base_url = os.getenv('MANIFEST_BASE_URL')
    
    # Fallback automático para URL se repositório for conhecido
    if not manifest_base_url and repo:
        parts = repo.split("/")
        if len(parts) == 2:
            manifest_base_url = f"https://{parts[0]}.github.io/{parts[1]}"

    try:
        data = None
        source = None

        # 1. Tentar ler do arquivo local (Se configurado)
        if manifest_file:
            if os.path.exists(manifest_file):
                try:
                    with open(manifest_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        source = f"file:{manifest_file}"
                except json.JSONDecodeError:
                    current_app.logger.exception(f"Erro de JSON inválido em MANIFEST_FILE {manifest_file}")
                    return jsonify({"error": "manifest_invalid_json", "details": f"Arquivo {manifest_file} não é um JSON válido"}), 500
                except Exception:
                    current_app.logger.exception(f"Erro ao ler MANIFEST_FILE {manifest_file}")
            else:
                current_app.logger.warning(f"MANIFEST_FILE configurado mas não encontrado: {manifest_file}")

        # 2. Tentar baixar da URL remota (Fonte principal ou Fallback)
        if not data and manifest_base_url:
            manifest_url = f"{manifest_base_url}/{env}.json"
            source = manifest_url
            try:
                response = requests.get(manifest_url, timeout=5)
                if response.status_code == 200:
                    data = response.json()
                else:
                    current_app.logger.error(f"Manifest não encontrado na URL {manifest_url}. Status: {response.status_code}")
            except json.JSONDecodeError:
                current_app.logger.exception(f"Resposta remota não é um JSON válido: {manifest_url}")
                return jsonify({"error": "manifest_invalid_json", "details": "A resposta da URL remota não é um JSON válido"}), 500
            except Exception:
                current_app.logger.exception(f"Erro ao baixar manifest de {manifest_url}")

        # Se não encontrou de nenhuma forma e não é dev puro
        if not data:
            if env == 'dev':
                return jsonify({
                    "current_version": __version__,
                    "latest_version": __version__,
                    "is_new": False,
                    "environment": env,
                    "source": "local_dev"
                })
            
            return jsonify({
                "error": "manifest_fetch_failed",
                "details": "Não foi possível obter o manifesto de nenhuma fonte configurada."
            }), 500

        # Extrair versão - Suporta 'version' ou 'latest_version'
        latest_version = data.get('version') or data.get('latest_version')
        
        if not latest_version:
            received_keys = list(data.keys())
            current_app.logger.error(f"Manifesto com formato inválido. Chaves recebidas: {received_keys}")
            return jsonify({
                "error": "manifest_invalid_format", 
                "details": f"Chave 'version' ou 'latest_version' não encontrada. Chaves recebidas: {received_keys}"
            }), 500

        # Comparação robusta SemVer
        current_v_tuple = parse_version(__version__)
        latest_v_tuple = parse_version(latest_version)
        is_new = latest_v_tuple > current_v_tuple
        
        return jsonify({
            "current_version": __version__,
            "latest_version": latest_version,
            "is_new": is_new,
            "maintenance": Configuracao.is_maintenance(),
            "tag": data.get('tag'),
            "commit": data.get('commit'),
            "date": data.get('date'),
            "environment": env,
            "source": source
        })
            
    except Exception:
        current_app.logger.exception("Erro fatal ao processar /api/system/latest")
        return jsonify({
            "error": "manifest_fetch_failed",
            "details": "Ocorreu um erro interno ao processar a verificação de versão."
        }), 500

@main_bp.route('/api/system/update/start', methods=['POST'])
@login_required
def system_update_start():
    if not current_user.is_admin:
        return jsonify({"error": "Unauthorized"}), 403
        
    if Configuracao.get_valor('UPDATE_IN_PROGRESS') == 'true':
        return jsonify({"error": "Update already in progress"}), 400
        
    try:
        # Pega a tag alvo do manifest antes de começar
        env = os.getenv('ENVIRONMENT', 'dev')
        repo = os.getenv('GITHUB_REPO', 'ademirapsantos/prp_financeiro')
        manifest_base_url = os.getenv('MANIFEST_BASE_URL')
        if not manifest_base_url and repo:
            parts = repo.split("/")
            if len(parts) == 2:
                manifest_base_url = f"https://{parts[0]}.github.io/{parts[1]}"
        
        manifest_url = f"{manifest_base_url}/{env}.json"
        res_m = requests.get(manifest_url, timeout=10)
        target_tag = 'latest'
        if res_m.status_code == 200:
            target_tag = res_m.json().get('tag', 'latest')

        # Configura estado inicial do update
        Configuracao.set_valor('UPDATE_IN_PROGRESS', 'true')
        Configuracao.set_valor('MAINTENANCE_MODE', 'true')
        Configuracao.set_valor('UPDATE_STARTED_AT', datetime.utcnow().isoformat())
        Configuracao.set_valor('UPDATE_LAST_ERROR', '')
        Configuracao.set_valor('UPDATE_TARGET_TAG', target_tag)
        # Tenta descobrir a tag atual (simplificado para o log)
        prev_tag = os.getenv(f'PRP_IMAGE_{env.upper()}', '').split(':')[-1] or 'v' + __version__
        Configuracao.set_valor('UPDATE_PREV_TAG', prev_tag)
        
        # Updater interaction is backend-only
        updater_base_url = os.getenv('UPDATER_BASE_URL')
        if not updater_base_url:
            if env == 'hml':
                updater_base_url = "http://prp-updater-hml:5005"
            elif env == 'prod':
                updater_base_url = "http://prp-updater-prod:5005"
            else:
                updater_base_url = "http://prp-updater:5005"
        
        # O updater.py do sidecar ouve em /api/update
        updater_url = f"{updater_base_url.rstrip('/')}/api/update"
        token = os.getenv('UPDATE_TOKEN', 'change_me_token')
        
        headers = {"Authorization": f"Bearer {token}"}
        
        # Call the sidecar (assíncrono para não prender o gunicorn por 300s)
        # Mas o updater atual é síncrono. Vamos tentar com timeout maior ou disparar em thread.
        # Como o objetivo é ser robusto, vamos aguardar um pouco.
        try:
            # Aumentamos o timeout pois o updater faz pull e recreate
            response = requests.post(updater_url, headers=headers, timeout=300)
            
            if response.status_code == 200:
                # Se for síncrono e retornar 200, já concluiu com sucesso
                Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
                Configuracao.set_valor('MAINTENANCE_MODE', 'false')
                return jsonify({"status": "success", "message": "Update concluído com sucesso"})
            else:
                # Falhou
                error_details = response.text
                Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
                Configuracao.set_valor('MAINTENANCE_MODE', 'false') # Mesmo falhando, sai da manutenção se der erro no updater
                Configuracao.set_valor('UPDATE_LAST_ERROR', error_details)
                return jsonify({"error": "updater_failed", "details": error_details}), response.status_code
        except requests.exceptions.Timeout:
            # Se der timeout, talvez ainda esteja processando. Mantemos as flags e tela de manutenção cuida.
            return jsonify({"status": "pending", "message": "Update em processamento (timeout na resposta)"})
        except requests.exceptions.ConnectionError:
            Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
            Configuracao.set_valor('MAINTENANCE_MODE', 'false')
            Configuracao.set_valor('UPDATE_LAST_ERROR', 'Falha ao conectar ao serviço updater')
            return jsonify({"error": "updater_unreachable", "details": "Falha ao conectar ao serviço updater (DNS/Rede)"}), 502
            
    except Exception as e:
        Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
        Configuracao.set_valor('MAINTENANCE_MODE', 'false')
        Configuracao.set_valor('UPDATE_LAST_ERROR', str(e))
        current_app.logger.exception("Erro ao acionar updater")
        return jsonify({"error": str(e)}), 500

@main_bp.route('/api/system/update/status')
def api_system_update_status():
    """Retorna o progresso e estado do update."""
    in_progress = Configuracao.get_valor('UPDATE_IN_PROGRESS') == 'true'
    maintenance = Configuracao.get_valor('MAINTENANCE_MODE') == 'true'
    started_at = Configuracao.get_valor('UPDATE_STARTED_AT')
    last_error = Configuracao.get_valor('UPDATE_LAST_ERROR')
    target_tag = Configuracao.get_valor('UPDATE_TARGET_TAG')
    prev_tag = Configuracao.get_valor('UPDATE_PREV_TAG')

    # Evita lock infinito: se passar do limite, libera manutenção e marca erro.
    if in_progress and started_at:
        try:
            started_dt = datetime.fromisoformat(started_at)
            max_minutes = int(os.getenv('UPDATE_MAX_MINUTES', '30'))
            if datetime.utcnow() - started_dt > timedelta(minutes=max_minutes):
                timeout_msg = f"Update excedeu tempo máximo de {max_minutes} minutos e foi finalizado automaticamente."
                Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
                Configuracao.set_valor('MAINTENANCE_MODE', 'false')
                Configuracao.set_valor('UPDATE_LAST_ERROR', timeout_msg)
                in_progress = False
                maintenance = False
                last_error = timeout_msg
        except Exception:
            current_app.logger.exception("Falha ao validar timeout de update")

    return jsonify({
        "in_progress": in_progress,
        "maintenance": maintenance,
        "started_at": started_at,
        "last_error": last_error,
        "target_tag": target_tag,
        "prev_tag": prev_tag,
        "current_version": __version__
    })

@main_bp.route('/api/system/maintenance/off-token', methods=['POST'])
def maintenance_off_token():
    """Desliga manutenção via token (Emergency/Updater usage)."""
    token = request.headers.get('Authorization')
    expected = f"Bearer {os.getenv('UPDATE_TOKEN', 'change_me_token')}"
    if token != expected:
        return jsonify({"error": "Unauthorized"}), 401
    
    Configuracao.set_valor('MAINTENANCE_MODE', 'false')
    Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
    return jsonify({"status": "success", "message": "Manutenção desligada via token"})

@main_bp.route('/api/system/update/finalize-token', methods=['POST'])
def api_system_update_finalize_token():
    """Finaliza update por token com status final informado pelo updater."""
    token = request.headers.get('Authorization')
    expected = f"Bearer {os.getenv('UPDATE_TOKEN', 'change_me_token')}"
    if token != expected:
        return jsonify({"error": "Unauthorized"}), 401

    payload = request.get_json(silent=True) or {}
    status = str(payload.get('status', 'success')).lower()
    error_message = (payload.get('error') or '').strip()

    Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
    Configuracao.set_valor('MAINTENANCE_MODE', 'false')
    Configuracao.set_valor('UPDATE_FINISHED_AT', datetime.utcnow().isoformat())

    if status in ('failed', 'error'):
        Configuracao.set_valor('UPDATE_LAST_ERROR', error_message or 'Falha na atualização.')
    else:
        Configuracao.set_valor('UPDATE_LAST_ERROR', '')

    return jsonify({"status": "success", "message": "Update finalizado por token"})

@main_bp.route('/api/notifications')
@login_required
def get_notifications():
    notificacoes = Notificacao.query.filter(
        (Notificacao.user_id == current_user.id) | (Notificacao.user_id == None)
    ).filter_by(lida=False).order_by(Notificacao.criada_em.desc()).all()
    
    return jsonify({
        "notifications": [{
            "id": n.id,
            "tipo": n.tipo,
            "titulo": n.titulo,
            "mensagem": n.mensagem,
            "payload_json": json.loads(n.payload_json) if n.payload_json else {},
            "criada_em": n.criada_em.isoformat()
        } for n in notificacoes]
    })

@main_bp.route('/api/notifications/<int:id>/read', methods=['POST'])
@login_required
def mark_notification_read(id):
    notif = Notificacao.query.get_or_404(id)
    if notif.user_id and notif.user_id != current_user.id:
        return jsonify({"error": "Unauthorized"}), 403
    
    notif.lida = True
    db.session.commit()
    return jsonify({"status": "success"})

@main_bp.route('/api/notifications', methods=['POST'])
@login_required
def create_notification_api():
    data = request.get_json()
    tipo = data.get('tipo', 'INFO')
    payload = data.get('payload', {})
    
    # Prevenção de duplicidade para UPDATE_AVAILABLE
    if tipo == 'UPDATE_AVAILABLE' and payload.get('latest_version'):
        versao = payload.get('latest_version')
        existente = Notificacao.query.filter_by(
            user_id=current_user.id, 
            tipo='UPDATE_AVAILABLE',
            lida=False
        ).all()
        
        for n in existente:
            try:
                p = json.loads(n.payload_json) if n.payload_json else {}
                if p.get('latest_version') == versao:
                    # Atualiza o payload se for mais recente (opcional, mantendo o existente)
                    return jsonify({"status": "success", "id": n.id, "message": "Notification already exists"})
            except (json.JSONDecodeError, TypeError):
                continue

    new_notif = Notificacao(
        user_id=current_user.id,
        tipo=tipo,
        titulo=data.get('titulo'),
        mensagem=data.get('mensagem'),
        payload_json=json.dumps(payload) if payload else None
    )
    db.session.add(new_notif)
    db.session.commit()
    return jsonify({"status": "success", "id": new_notif.id})

@main_bp.route('/api/system/maintenance/on', methods=['POST'])
@login_required
def maintenance_on():
    if not current_user.is_admin:
        return jsonify({"error": "Unauthorized"}), 403
    Configuracao.set_valor('MAINTENANCE_MODE', 'true')
    return jsonify({"status": "success"})

@main_bp.route('/api/system/maintenance/off', methods=['POST'])
@login_required
def maintenance_off():
    if not current_user.is_admin:
        return jsonify({"error": "Unauthorized"}), 403
    Configuracao.set_valor('MAINTENANCE_MODE', 'false')
    Configuracao.set_valor('UPDATE_IN_PROGRESS', 'false')
    return jsonify({"status": "success"})

